import pandas as pd  
from fpdf import FPDF  # Geração de arquivos PDF
from geopy.distance import geodesic  # Cálculo de distância geodésica
from geopy.geocoders import Nominatim  # Geocodificador Nominatim (OSM)
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import numpy as np  # Operações numéricas e matrizes
import os  # arquivos
import unicodedata  # Remoção de acentos/normalização de texto
import json  
from datetime import datetime, timedelta  
from tqdm import tqdm  
import colorama
from colorama import Fore, Style  
import requests  # Requisições HTTP
import openpyxl  
from openpyxl.styles import PatternFill  # Estilo de células (preenchimento)
import time  # Pausas entre requisições e backoff
import re  # Expressões regulares

colorama.init()  # Habilita cores no terminal no Windows

# Configuração do geocodificador Nominatim
geolocator = Nominatim(user_agent="juninho.junirj@gmail.com")

def print_colorido(texto, cor=Fore.WHITE, estilo=Style.NORMAL):  
    print(f"{estilo}{cor}{texto}{Style.RESET_ALL}")  


# Função para detectar se o endereço está no formato de coordenadas
def is_coordenada(texto):  
    if not isinstance(texto, str):  
        return False
    padrao = r"^\s*(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)(?:\s*[,;]\s*.*)?$"  # Captura dois decimais separados por vírgula
    return re.match(padrao, texto) is not None  

def extrair_coordenada(texto):  # Extrai (lat, lon) de uma string
    # Regex para extrair apenas latitude e longitude, ignorando texto adicional
    padrao = r"^\s*(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)"  # Usa grupos para capturar lat e lon
    m = re.match(padrao, texto)  
    if m:  
        return (float(m.group(1)), float(m.group(2)))  # Converte para float e retorna dupla
    return None  

def remover_acentos(texto):  
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                  if unicodedata.category(c) != 'Mn')  

def limpar_endereco(endereco):
    """Limpa e padroniza o endereço para melhorar a geocodificação."""
    if not isinstance(endereco, str):
        return ""
    # Remove espaços extras no início e no fim
    endereco = endereco.strip()
    # Remove caracteres especiais, exceto os essenciais (vírgula, hífen, número)
    endereco = re.sub(r'[^\w\s,-]', '', endereco)
    # Substitui múltiplos espaços por um único
    endereco = re.sub(r'\s+', ' ', endereco)
    return endereco

def enriquecer_endereco(endereco, cidade):
    """Padroniza o endereço para o formato amigável ao OpenStreetMap."""
    if not cidade:
        return endereco

    endereco_str = str(endereco).strip()

    # Substitui " - SP" ou "-SP" por ", SP"
    endereco_str = re.sub(r'\s*-\s*([A-Z]{2})\b', r', \1', endereco_str, flags=re.IGNORECASE)

    endereco_norm = remover_acentos(endereco_str).lower()
    cidade_norm = remover_acentos(cidade).lower()

    # Se a cidade não estiver na string, adiciona
    if cidade_norm not in endereco_norm:
        endereco_str = f"{endereco_str}, {cidade}"

    # Adiciona "Brasil" no final se não existir, ajuda MUITO o Nominatim
    if "brasil" not in endereco_norm:
        endereco_str = f"{endereco_str}, Brasil"

    return endereco_str

# Cache para geocodificação com timestamp
CACHE_FILE = "geocodificacao_cache.json"  # Arquivo onde o cache de geocodificação é salvo
CACHE_EXPIRATION_DAYS = 30  # Cache expira após 30 dias

def carregar_cache():  # Lê o cache de geocodificação do disco, removendo entradas expiradas
    if os.path.exists(CACHE_FILE):  # Verifica se o arquivo existe
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:  # Abre arquivo JSON
                cache_data = json.load(f)  # Carrega conteúdo
                current_time = datetime.now()  # Hora atual
                cache_data = {  # Filtra apenas entradas não expiradas
                    k: v for k, v in cache_data.items()
                    if datetime.fromisoformat(v['timestamp']) + timedelta(days=CACHE_EXPIRATION_DAYS) > current_time
                }
                return cache_data  # Retorna cache limpo
        except Exception as e:  # Qualquer erro de leitura/parse
            print_colorido(f"Erro ao carregar cache: {str(e)}", Fore.RED)  
            return {}
    return {}

def salvar_cache(cache):  # Persiste cache de geocodificação em disco
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:  
            json.dump(cache, f, ensure_ascii=False, indent=2)  # Salva JSON
    except Exception as e:  
        print_colorido(f"Erro ao salvar cache: {str(e)}", Fore.RED)

def geocodificar_endereco_nominatim(endereco, max_tentativas=3, intervalo=2):
    """Converte um endereço em coordenadas (latitude, longitude) usando Nominatim, com tratamento de erro detalhado."""
    motivo = "Falha na geocodificação"  # Mensagem de erro padrão
    for tentativa in range(max_tentativas):
        try:
            time.sleep(1.1) # Garante que nunca bata mais de 1 req/seg
            location = geolocator.geocode(endereco, timeout=15)
            if location:
                return {'coords': (location.latitude, location.longitude), 'timestamp': datetime.now().isoformat()}
            else:
                return {'error': 'Endereço não encontrado'}
        except GeocoderTimedOut:
            motivo = "Timeout do serviço"
            print_colorido(f"{motivo} ao geocodificar '{endereco}' (tentativa {tentativa + 1}/{max_tentativas})", Fore.YELLOW)
        except GeocoderServiceError as e:
            if "429" in str(e):
                print_colorido("Limite atingido! Dormindo por 10 segundos...", Fore.RED)
                time.sleep(10) # Pausa longa se for bloqueado
                return {'error': "429: Limite excedido"}
            return {'error': f"Erro de serviço: {e}"} # Faltou retornar algo se não for 429
        except Exception as e:
            motivo = f"Erro desconhecido: {e}"
            print_colorido(f"Erro inesperado ao geocodificar '{endereco}': {e}", Fore.RED)

        if tentativa < max_tentativas - 1:
            time.sleep(intervalo)
        else:
            return {'error': motivo}
    return {'error': motivo}

def geocodificar_endereco_photon(endereco, max_tentativas=3, intervalo=5):
    params = {'q': endereco, 'limit': 1}
    headers = {'User-Agent': 'juninho.junirj@gmail.com'} # Adicione isso
    motivo = "Falha na geocodificação (Photon)"
    for tentativa in range(max_tentativas):
        try:
            response = requests.get("https://photon.komoot.io/api/", params=params, headers=headers, timeout=15)
            response.raise_for_status()
            data = response.json()
            if data and 'features' in data and data['features']:
                lon, lat = data['features'][0]['geometry']['coordinates']
                return {'coords': (lat, lon), 'timestamp': datetime.now().isoformat()}
            else:
                return {'error': 'Endereço não encontrado (Photon)'}
        except requests.exceptions.Timeout:
            motivo = "Timeout do serviço (Photon)"
            print_colorido(f"{motivo} para '{endereco}' (tentativa {tentativa + 1}/{max_tentativas})", Fore.YELLOW)
        except requests.exceptions.RequestException as e:
            motivo = f"Erro de requisição (Photon): {e}"
            print_colorido(f"Erro na requisição para Photon: {e}", Fore.RED)
            return {'error': motivo}

        if tentativa < max_tentativas - 1:
            time.sleep(intervalo)
        else:
            return {'error': motivo}
    return {'error': motivo}

def geocodificar_endereco(endereco, max_tentativas=3, intervalo=5):
    """
    Tenta geocodificar um endereço usando Nominatim e, se falhar,
    usa a API Photon como fallback.
    """
    # 1. Tentar com Nominatim
    resultado_nominatim = geocodificar_endereco_nominatim(endereco, max_tentativas=2, intervalo=intervalo)
    if resultado_nominatim and 'coords' in resultado_nominatim:
        resultado_nominatim['provider'] = 'Nominatim'
        return resultado_nominatim

    motivo_falha_nominatim = resultado_nominatim.get('error', 'Falha') if resultado_nominatim else 'Falha'
    print_colorido(f"Falha no Nominatim para '{endereco}' ({motivo_falha_nominatim}). Tentando fallback...", Fore.YELLOW)

    # 2. Tentar com Photon
    resultado_photon = geocodificar_endereco_photon(endereco, max_tentativas=max_tentativas, intervalo=1)
    if resultado_photon and 'coords' in resultado_photon:
        resultado_photon['provider'] = 'Photon'
        return resultado_photon

    motivo_falha_photon = resultado_photon.get('error', 'Falha') if resultado_photon else 'Falha'
    return {'error': f"Nominatim: {motivo_falha_nominatim} | Photon: {motivo_falha_photon}"}

def calcular_distancia_rua(coords1, coords2):  # Fallback: distância geodésica (reta) entre coordenadas
    try:
        lat1, lon1 = float(coords1[0]), float(coords1[1])  # Normaliza para float
        lat2, lon2 = float(coords2[0]), float(coords2[1])
        if not (-90 <= lat1 <= 90) or not (-90 <= lat2 <= 90) or \
           not (-180 <= lon1 <= 180) or not (-180 <= lon2 <= 180):  # Validação básica
            print_colorido(f"Coordenadas inválidas: ({lat1}, {lon1}) ou ({lat2}, {lon2})", Fore.RED)
            return float('inf')  # Usa infinito para sinalizar distância inválida
        dist = geodesic((lat1, lon1), (lat2, lon2)).kilometers  # Distância geodésica em km
        if dist > 500:  # Proteção contra outliers improváveis
            print_colorido(f"Distância suspeita: {dist:.2f}km entre ({lat1}, {lon1}) e ({lat2}, {lon2})", Fore.YELLOW)
            return float('inf')  # Ignora distâncias suspeitas
        return dist  # Distância válida
    except Exception as e:  # Tratamento de erro genérico
        print_colorido(f"Erro ao calcular distância: {str(e)}", Fore.RED)
        return float('inf')  # Considera inválido

def calcular_matriz_distancia(coordenadas):
    """Calcula a matriz de distância (km) e duração (min) usando distância geodésica."""
    n = len(coordenadas)
    dist_matrix = np.full((n, n), float('inf'))
    dur_matrix = np.full((n, n), float('inf'))

    with tqdm(total=n*n, desc="Calculando Matriz de Distância/Duração (Geodésica)") as pbar:
        for i in range(n):
            for j in range(n):
                if i == j:
                    dist_matrix[i][j] = 0
                    dur_matrix[i][j] = 0
                else:
                    # Fator de 1.4 para simular distância de rua
                    dist = calcular_distancia_rua(coordenadas[i], coordenadas[j]) * 1.4
                    dist_matrix[i][j] = dist
                    # Duração estimada com velocidade média de 40 km/h
                    if dist != float('inf'):
                        dur_matrix[i][j] = (dist / 40.0) * 60.0
                pbar.update(1)

    return dist_matrix, dur_matrix

def identificar_outliers(dist_matrix, enderecos_validos, limite_desvio=2.5, p80_limite_km=300):  # Detecta pontos muito afastados
    """Identifica pontos que estão muito distantes da média"""
    n = len(dist_matrix)  # Número de pontos
    if n <= 1:  # Com 0 ou 1 ponto não há outliers
        return [], []

    # Calcula a média e desvio padrão das distâncias
    distancias = []  # Coleta todas as distâncias válidas
    for i in range(n):
        for j in range(n):
            if i != j and dist_matrix[i][j] != float('inf'):
                distancias.append(dist_matrix[i][j])  # Adiciona distância i->j
    
    if not distancias:  # Se não há distâncias válidas
        return [], []  # Nada a analisar

    media = sum(distancias) / len(distancias)  # Média das distâncias
    desvio = (sum((x - media) ** 2 for x in distancias) / len(distancias)) ** 0.5  # Desvio padrão
    
    # Identifica outliers
    outliers = []  # Índices considerados outliers
    pontos_principais = []  # Demais pontos
    
    for i in range(n):  # Para cada ponto, avalia sua distância média aos outros
        distancias_ponto = [dist_matrix[i][j] for j in range(n) if i != j and dist_matrix[i][j] != float('inf')]
        if not distancias_ponto:  # Se não há vizinhos válidos
            continue
        
        media_ponto = sum(distancias_ponto) / len(distancias_ponto)  # Média local
        # Calcula percentil 80% das distâncias do ponto
        ordenadas = sorted(distancias_ponto)
        idx_p80 = max(0, min(len(ordenadas) - 1, int(0.8 * (len(ordenadas) - 1))))
        p80 = ordenadas[idx_p80]
        # Critérios: z-score alto OU percentil 80 muito alto (a maioria distante)
        if (media_ponto > media + (limite_desvio * desvio)) or (p80 > p80_limite_km):
            outliers.append(i)  # Marca como outlier
            print_colorido(f"Ponto identificado como outlier: {enderecos_validos[i]} (distância média: {media_ponto:.2f} km)", Fore.YELLOW)
        else:
            pontos_principais.append(i)  # Ponto normal
    
    return pontos_principais, outliers  # Retorna listas de índices

def encontrar_melhor_rota(dist_matrix, enderecos_validos):
    n = len(dist_matrix)
    if n <= 1:
        return [0] if n == 1 else []

    # 1. Construir rota inicial com o vizinho mais próximo
    print_colorido("Construindo rota inicial (Vizinho Mais Próximo)...", Fore.CYAN)
    ponto_partida = 0
    rota_inicial = [ponto_partida]
    nao_visitados = set(range(n))
    nao_visitados.remove(ponto_partida)
    
    ponto_atual = ponto_partida
    while nao_visitados:
        proximo_ponto = min(nao_visitados, key=lambda ponto: dist_matrix[ponto_atual][ponto])
        distancia = dist_matrix[ponto_atual][proximo_ponto]
        print_colorido(f"  Adicionando trecho: {enderecos_validos[ponto_atual]} -> {enderecos_validos[proximo_ponto]} ({distancia:.2f} km)", Fore.WHITE)
        rota_inicial.append(proximo_ponto)
        nao_visitados.remove(proximo_ponto)
        ponto_atual = proximo_ponto

    # 2. Otimizar a rota com 2-opt
    print_colorido("\nOtimizando a rota com 2-opt...", Fore.YELLOW)
    rota_otimizada = rota_inicial
    melhorou = True
    tentativas = 0
    while melhorou and tentativas < 1000: # Adiciona um limite de tentativas para evitar loops infinitos
        melhorou = False
        for i in range(1, n - 2):
            for j in range(i + 1, n -1):
                # Segmento da rota a ser revertido é de i a j
                # Arestas atuais: (i-1) -> (i) e (j) -> (j+1)
                # Novas arestas: (i-1) -> (j) e (i) -> (j+1)

                ponto_a, ponto_b = rota_otimizada[i-1], rota_otimizada[i]
                ponto_c, ponto_d = rota_otimizada[j], rota_otimizada[j+1]

                dist_atual = dist_matrix[ponto_a][ponto_b] + dist_matrix[ponto_c][ponto_d]
                dist_nova = dist_matrix[ponto_a][ponto_c] + dist_matrix[ponto_b][ponto_d]

                if dist_nova < dist_atual:
                    # Inverte o segmento da rota
                    segmento_invertido = rota_otimizada[i:j+1]
                    segmento_invertido.reverse()
                    rota_otimizada[i:j+1] = segmento_invertido

                    melhorou = True
                    ganho = dist_atual - dist_nova
                    print_colorido(f"  Otimização 2-opt: Trocando arestas ({i-1}-{i}, {j}-{j+1}). Ganho: {ganho:.2f} km", Fore.GREEN)
                    # Break para reiniciar a busca a partir do início com a nova rota
                    break
            if melhorou:
                break
        tentativas += 1

    if tentativas >= 1000:
        print_colorido("Atingido o limite de iterações do 2-opt.", Fore.YELLOW)

    # Calcula a distância total da rota otimizada
    distancia_total = 0
    for i in range(len(rota_otimizada) - 1):
        distancia_total += dist_matrix[rota_otimizada[i]][rota_otimizada[i + 1]]
    
    print_colorido(f"\nDistância total da rota otimizada: {distancia_total:.2f} km", Fore.GREEN)
    return rota_otimizada

# CARREGAR CONFIGURAÇÕES
def carregar_config():
    try:
        with open("config.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print_colorido("❌ Erro: O arquivo de configuração 'config.json' não foi encontrado.", Fore.RED)
        exit(1)
    except json.JSONDecodeError:
        print_colorido("❌ Erro: O arquivo 'config.json' contém um JSON inválido.", Fore.RED)
        exit(1)

def main():
    """Função principal que orquestra todo o processo de otimização de rota."""
    config = carregar_config()
    arquivo_excel = config.get("arquivo_excel", "ENDERECOS-ROTA.xlsx")
    nome_coluna_enderecos = config.get("nome_coluna_enderecos", "Endereco")
    nome_coluna_nomes = config.get("nome_coluna_nomes", "Nome")
    ponto_partida_bruto = config.get("ponto_partida", "Rua Floriano Peixoto, 368, Centro, Itapuí - SP")
    ponto_partida = ponto_partida_bruto if is_coordenada(ponto_partida_bruto) else limpar_endereco(ponto_partida_bruto)

    import sys

    if len(sys.argv) > 1:
        cidade = sys.argv[1]
    else:
        cidade = input("Digite a cidade das entregas: ").strip()

    try:
        print_colorido("\n🚀 Iniciando processamento...", Fore.GREEN, Style.BRIGHT)

        if not os.path.exists(arquivo_excel):
            print_colorido(f"❌ Erro: O arquivo {arquivo_excel} não foi encontrado.", Fore.RED)
            exit(1)

        print_colorido("\n📊 Lendo planilha...", Fore.CYAN)
        try:
            df = pd.read_excel(arquivo_excel)
            df = df[df[nome_coluna_nomes].notna() & (df[nome_coluna_nomes] != "")]
            enderecos_brutos = df[nome_coluna_enderecos].dropna().tolist()
            enderecos = [end if is_coordenada(end) else limpar_endereco(end) for end in enderecos_brutos]
            nomes = df[nome_coluna_nomes].fillna("").tolist()
            print_colorido(f"✅ Total de endereços encontrados: {len(enderecos)}", Fore.GREEN)
        except Exception as e:
            print_colorido(f"❌ Erro ao ler planilha: {str(e)}", Fore.RED)
            exit(1)

        if not enderecos:
            print_colorido("❌ Erro: Nenhum endereço encontrado na planilha.", Fore.RED)
            exit(1)

        print_colorido("\n🌍 Iniciando geocodificação...", Fore.CYAN)
        coordenadas = []
        enderecos_validos = []
        enderecos_com_erro = []
        cache = carregar_cache()

        ponto_partida_enriquecido = enriquecer_endereco(ponto_partida, cidade)
        print_colorido(f"\n📍 Processando ponto de partida: {ponto_partida_enriquecido}", Fore.CYAN)

        if is_coordenada(ponto_partida):
            coords = extrair_coordenada(ponto_partida)
            if coords:
                coordenadas.append(coords)
                enderecos_validos.append(ponto_partida)
            else:
                print_colorido(f"❌ Erro: Não foi possível interpretar as coordenadas do ponto de partida: {ponto_partida}", Fore.RED)
                exit(1)
        elif ponto_partida in cache:
            print_colorido("✅ Usando coordenadas do cache para ponto de partida", Fore.GREEN)
            coordenadas.append(tuple(cache[ponto_partida]['coords']))
            enderecos_validos.append(ponto_partida)
        else:
            resultado = geocodificar_endereco(ponto_partida_enriquecido)
            if resultado and 'coords' in resultado:
                coordenadas.append(resultado['coords'])
                enderecos_validos.append(ponto_partida)
                cache[ponto_partida] = resultado
                provedor = resultado.get('provider', 'desconhecido')
                print_colorido(f"✅ Ponto de partida geocodificado com {provedor}", Fore.GREEN)
            else:
                motivo_erro = resultado.get('error', 'desconhecido') if resultado else 'desconhecido'
                print_colorido(f"❌ Erro: Não foi possível geocodificar o ponto de partida: {ponto_partida} (Motivo: {motivo_erro})", Fore.RED)
                exit(1)

        def processar_endereco(endereco, cidade, cache): # <-- Removido o parâmetro 'lock'
            endereco_enriquecido = enriquecer_endereco(endereco, cidade)
            if is_coordenada(endereco):
                coords = extrair_coordenada(endereco)
                return (endereco, coords, 'coordenada', None) if coords else (endereco, None, 'erro', 'Coordenada inválida')
            
            if endereco in cache:
                provedor = cache[endereco].get('provider', 'desconhecido')
                return (endereco, tuple(cache[endereco]['coords']), f'cache ({provedor})', None)
            
            resultado = geocodificar_endereco(endereco_enriquecido)
            
            if resultado and 'coords' in resultado:
                cache[endereco] = resultado # <-- Salva direto no dicionário, sem precisar de 'with lock:'
                salvar_cache(cache)         # <-- Salva o arquivo JSON em tempo real
                status = f"geocodificado ({resultado.get('provider', 'desconhecido')})"
                return (endereco, resultado['coords'], status, None)
            else:
                motivo_erro = resultado.get('error', 'Erro desconhecido') if resultado else 'Erro desconhecido'
                return (endereco, None, 'erro', motivo_erro)

        print_colorido("\n🔄 Geocodificando endereços...", Fore.CYAN)
        resultados = []
        
        # Loop sequencial normal, já que API gratuita não permite paralelismo
        for end in tqdm(enderecos, desc="Progresso", unit="endereço"):
            # Chamada limpa, passando apenas o endereço, cidade e o cache
            resultado = processar_endereco(end, cidade=cidade, cache=cache)
            resultados.append(resultado)

        salvar_cache(cache) # Salva uma última vez ao final do loop por segurança

        for i, (endereco, coords, status, motivo_erro) in enumerate(resultados, 1):
            if coords:
                coordenadas.append(tuple(coords))
                enderecos_validos.append(endereco)
            else:
                enderecos_com_erro.append((i, endereco, motivo_erro))
            if 'cache' in status:
                print_colorido(f"Usando {status} para: {endereco}", Fore.YELLOW)
            elif 'geocodificado' in status:
                provedor = status.split('(')[-1].replace(')', '')
                cor = Fore.CYAN if 'Photon' in provedor else Fore.GREEN
                print_colorido(f"Geocodificado com {provedor}: {endereco}", cor)
            elif status == 'coordenada':
                print_colorido(f"Endereço já é coordenada: {endereco}", Fore.CYAN)
            else:
                print_colorido(f"Erro ao geocodificar: {endereco} (Motivo: {motivo_erro})", Fore.RED)

        def marcar_enderecos_erro_excel(arquivo_excel, enderecos_com_erro):
            try:
                wb = openpyxl.load_workbook(arquivo_excel)
                ws = wb.active
                num_cols = ws.max_column
                fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                for linha, _, _ in enderecos_com_erro:
                    for col in range(1, num_cols + 1):
                        ws.cell(row=linha + 1, column=col).fill = fill
                wb.save(arquivo_excel)
                print_colorido("Linhas dos endereços com erro marcadas em vermelho na planilha.", Fore.RED)
            except Exception as e:
                print_colorido(f"Erro ao marcar células na planilha: {str(e)}", Fore.RED)

        if enderecos_com_erro:
            marcar_enderecos_erro_excel(arquivo_excel, enderecos_com_erro)

        if len(enderecos_validos) <= 1:
            print_colorido("❌ Erro: Nenhum endereço foi geocodificado com sucesso além do ponto de partida.", Fore.RED)
            exit(1)

        print_colorido(f"\n✅ Total de endereços geocodificados com sucesso: {len(enderecos_validos)}", Fore.GREEN)
        print_colorido(f"⚠️ Total de endereços com erro: {len(enderecos_com_erro)}", Fore.YELLOW)

        print_colorido("\n📏 Calculando matriz de distância...", Fore.CYAN)
        dist_matrix, dur_matrix = calcular_matriz_distancia(coordenadas)

        # Validação da Matriz de Distância
        if np.isinf(dist_matrix).any():
            print_colorido("❌ Erro: A matriz de distância contém valores inválidos (infinitos).", Fore.RED)
            print_colorido("   Isso geralmente ocorre por problemas na geocodificação ou coordenadas inválidas.", Fore.YELLOW)
            print_colorido("   Verifique os endereços marcados com erro na planilha e tente novamente.", Fore.YELLOW)

            problematic_pairs = []
            for i in range(len(dist_matrix)):
                for j in range(i + 1, len(dist_matrix)):
                    if np.isinf(dist_matrix[i][j]):
                        problematic_pairs.append((enderecos_validos[i], enderecos_validos[j]))

            if problematic_pairs:
                print_colorido("   Pares de endereços com distância infinita (amostra):", Fore.YELLOW)
                for end1, end2 in problematic_pairs[:5]:
                    print_colorido(f"     - De '{end1}' para '{end2}'", Fore.WHITE)
            exit(1)

        for i in range(len(coordenadas)):
            for j in range(i + 1, len(coordenadas)):
                dist = dist_matrix[i][j]
                if dist != float('inf'):
                    print_colorido(f"   De {enderecos_validos[i]} para {enderecos_validos[j]}: {dist:.2f} km", Fore.WHITE)

        pontos_principais, outliers_idx = identificar_outliers(dist_matrix, enderecos_validos)
        outliers_info = []
        for idx in outliers_idx:
            distancias_ponto = [dist_matrix[idx][j] for j in range(len(dist_matrix)) if idx != j and dist_matrix[idx][j] != float('inf')]
            media_ponto = (sum(distancias_ponto) / len(distancias_ponto)) if distancias_ponto else 0
            endereco_out = enderecos_validos[idx]
            nome_out = nomes[enderecos.index(endereco_out)] if endereco_out in enderecos else ""
            outliers_info.append((idx, nome_out, endereco_out, media_ponto))
        outliers_idx_set = set(outliers_idx)

        print_colorido("\n🗺️ Calculando melhor rota...", Fore.CYAN)
        ordem_rota = encontrar_melhor_rota(dist_matrix, enderecos_validos)

        if ordem_rota is None or len(ordem_rota) != len(coordenadas):
            print_colorido("❌ Erro: A rota não pôde ser calculada ou não inclui todos os pontos!", Fore.RED)
            exit(1)

        distancia_total, tempo_total_min, distancias_parciais, duracoes_parciais = 0, 0, [], []
        for i in range(len(ordem_rota) - 1):
            origem, destino = ordem_rota[i], ordem_rota[i+1]
            dist, dur = dist_matrix[origem][destino], dur_matrix[origem][destino]
            distancia_total += dist
            tempo_total_min += dur
            distancias_parciais.append(dist)
            duracoes_parciais.append(dur)

        print_colorido(f"\n📊 Distância total da rota: {distancia_total:.2f} km", Fore.GREEN)
        print_colorido(f"⏱️ Tempo total estimado: {int(tempo_total_min // 60)}h {int(tempo_total_min % 60)}min", Fore.GREEN)

        enderecos_ordenados = [enderecos_validos[i] for i in ordem_rota]
        nomes_ordenados = [nomes[enderecos.index(e)] if e in enderecos else "" for e in enderecos_ordenados]
        links = [f"https://www.google.com/maps/dir/?api=1&destination={remover_acentos(e).replace(' ', '+')}" for e in enderecos_ordenados]

        print_colorido("\n📄 Gerando PDF...", Fore.CYAN)
        class PDF(FPDF):
            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 8)
                self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()

        if os.path.exists("./assets/logo.png"):
            pdf.image("./assets/logo.png", x=170, y=10, w=31.5)

        pdf.set_font("Arial", "B", 16)
        data_atual = datetime.now().strftime("%d/%m/%Y")
        titulo = f"Rota de Entregas - {cidade}"
        subtitulo = f"Gerado em: {data_atual}"

        nome_arquivo = remover_acentos(f"{datetime.now().strftime('%Y-%m-%d')} - Rota de Entregas - {cidade}").replace("/", "-")
        pasta_rotas = "ROTAS-GERADAS"
        if not os.path.exists(pasta_rotas):
            os.makedirs(pasta_rotas)
        arquivo_saida_pdf = os.path.join(pasta_rotas, f"{nome_arquivo}.pdf")

        pdf.cell(0, 10, titulo, ln=True, align="C")
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 6, subtitulo, ln=True, align="C")
        pdf.ln(8)

        def formatar_minutos(minutos):
            try:
                h = int(minutos // 60)
                m = int(round(minutos % 60))
                if m == 60: h += 1; m = 0
                return f"{h}h {m:02d}min"
            except: return "-"

        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, "Resumo da Rota", ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.multi_cell(0, 6,
            f"Ponto de Partida: {ponto_partida}\n"
            f"Distância Total: {distancia_total:.1f} km\n"
            f"Tempo Total de Viagem: {formatar_minutos(tempo_total_min)}\n"
            f"Número de Entregas: {len(enderecos_ordenados) - 1}",
            border=1, align="L"
        )
        pdf.ln(8)

        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, "Ordem de Visita", ln=True)

        pdf.set_fill_color(255, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 8)
        w_num, w_nome, w_end, w_dist, w_tempo, w_link = 8, 50, 66, 18, 18, 30
        pdf.cell(w_num, 8, "Nº", 1, 0, "C", True)
        pdf.cell(w_nome, 8, "Nome", 1, 0, "C", True)
        pdf.cell(w_end, 8, "Endereço", 1, 0, "C", True)
        pdf.cell(w_dist, 8, "Dist (km)", 1, 0, "C", True)
        pdf.cell(w_tempo, 8, "Tempo", 1, 0, "C", True)
        pdf.cell(w_link, 8, "Link (Maps)", 1, 1, "C", True)
        pdf.set_text_color(0, 0, 0)

        pdf.set_font("Arial", "", 8)
        cor_linha = 0
        for i, (nome, endereco, link) in enumerate(zip(nomes_ordenados[1:], enderecos_ordenados[1:], links[1:]), 1):
            dist_trecho, dur_trecho = distancias_parciais[i-1], duracoes_parciais[i-1]
            idx_original = ordem_rota[i]
            is_outlier = idx_original in outliers_idx_set
            pdf.set_fill_color(245, 245, 245) if cor_linha % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            if is_outlier: pdf.set_fill_color(255, 255, 204)

            x_inicial, y_inicial = pdf.get_x(), pdf.get_y()
            line_height, altura_max = 5, 5

            pdf.multi_cell(w_num, line_height, f"#{i}", border=0, align='C', fill=True)
            h = pdf.get_y() - y_inicial; altura_max = max(h, altura_max)
            pdf.set_xy(x_inicial + w_num, y_inicial)

            pdf.multi_cell(w_nome, line_height, str(nome), border=0, align='L', fill=True)
            h = pdf.get_y() - y_inicial; altura_max = max(h, altura_max)
            pdf.set_xy(x_inicial + w_num + w_nome, y_inicial)

            pdf.multi_cell(w_end, line_height, str(endereco), border=0, align='L', fill=True)
            h = pdf.get_y() - y_inicial; altura_max = max(h, altura_max)

            pdf.set_xy(x_inicial, y_inicial)
            pdf.cell(w_num, altura_max, "", border=1)
            pdf.cell(w_nome, altura_max, "", border=1)
            pdf.cell(w_end, altura_max, "", border=1)

            pdf.set_xy(x_inicial + w_num + w_nome + w_end, y_inicial)
            pdf.cell(w_dist, altura_max, f"{dist_trecho:.1f}", 1, 0, "C", True)
            pdf.cell(w_tempo, altura_max, f"{int(round(dur_trecho))} min", 1, 0, "C", True)

            pdf.set_text_color(0, 0, 255); pdf.set_font("", "U")
            pdf.cell(w_link, altura_max, "Abrir no Maps", 1, 1, "C", True, link=link)
            pdf.set_font(""); pdf.set_text_color(0, 0, 0)

            cor_linha += 1
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_fill_color(255, 0, 0); pdf.set_text_color(255, 255, 255); pdf.set_font("Arial", "B", 8)
                pdf.cell(w_num, 8, "Nº", 1, 0, "C", True); pdf.cell(w_nome, 8, "Nome", 1, 0, "C", True); pdf.cell(w_end, 8, "Endereço", 1, 0, "C", True)
                pdf.cell(w_dist, 8, "Dist (km)", 1, 0, "C", True); pdf.cell(w_tempo, 8, "Tempo", 1, 0, "C", True); pdf.cell(w_link, 8, "Link (Maps)", 1, 1, "C", True)
                pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 8)

        if enderecos_com_erro:
            if pdf.get_y() > 220: pdf.add_page()
            pdf.ln(10)
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Apêndice A: Endereços com Erro", ln=True, align="L")
            pdf.set_font("Arial", "", 10)
            pdf.multi_cell(0, 6, "A lista abaixo contém os endereços que não puderam ser encontrados. Verifique se há erros de digitação ou falta de informações. As linhas correspondentes no Excel foram marcadas em vermelho.", align="L")
            pdf.ln(3)
            pdf.set_fill_color(255, 0, 0); pdf.set_text_color(255, 255, 255); pdf.set_font("Arial", "B", 10)
            w_linha, w_nome, w_end, w_motivo = 15, 45, 80, 50
            pdf.cell(w_linha, 8, "Linha", 1, 0, "C", True); pdf.cell(w_nome, 8, "Nome", 1, 0, "C", True)
            pdf.cell(w_end, 8, "Endereço", 1, 0, "C", True); pdf.cell(w_motivo, 8, "Motivo do Erro", 1, 1, "C", True)
            pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 8)
            for linha, endereco, motivo in enderecos_com_erro:
                nome_cliente = nomes[linha - 1] if 0 <= linha - 1 < len(nomes) else ""
                pdf.cell(w_linha, 8, str(linha), 1, 0, "C")
                pdf.cell(w_nome, 8, str(nome_cliente)[:28], 1, 0, "L")
                pdf.cell(w_end, 8, str(endereco)[:50], 1, 0, "L")
                pdf.cell(w_motivo, 8, str(motivo)[:30], 1, 1, "L")

        if outliers_info:
            if pdf.get_y() > 220: pdf.add_page()
            pdf.ln(10)
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Apêndice B: Endereços Outliers", ln=True, align="L")
            pdf.set_font("Arial", "", 10)
            pdf.multi_cell(0, 6, "Os 'outliers' são pontos muito distantes do grupo. Podem ser erros de geocodificação ou entregas em cidades vizinhas. Verifique se as coordenadas estão corretas.", align="L")
            pdf.ln(3)
            pdf.set_fill_color(255, 0, 0); pdf.set_text_color(255, 255, 255); pdf.set_font("Arial", "B", 10)
            pdf.cell(20, 8, "Índice", 1, 0, "C", True); pdf.cell(50, 8, "Nome", 1, 0, "C", True)
            pdf.cell(80, 8, "Endereço", 1, 0, "C", True); pdf.cell(40, 8, "Dist. Média (km)", 1, 1, "C", True)
            pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 9)
            for idx, nome_out, endereco_out, media_ponto in outliers_info:
                pdf.cell(20, 8, str(idx), 1, 0, "C")
                pdf.cell(50, 8, str(nome_out)[:30], 1, 0, "L")
                pdf.cell(80, 8, str(endereco_out)[:50], 1, 0, "L")
                pdf.cell(40, 8, f"{media_ponto:.1f}", 1, 1, "C")

        pdf.output(arquivo_saida_pdf)
        print_colorido(f"\n✅ PDF gerado com sucesso: {arquivo_saida_pdf}", Fore.GREEN)

    except Exception as e:
        print_colorido(f"\n❌ Erro inesperado: {str(e)}", Fore.RED)
        import traceback
        print_colorido("Detalhes do erro:", Fore.RED)
        print_colorido(traceback.format_exc(), Fore.RED)
        exit(1)

if __name__ == "__main__":
    main()