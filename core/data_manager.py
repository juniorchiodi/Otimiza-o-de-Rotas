import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from utils.formatadores import is_coordenada, limpar_endereco
import re

def limpar_cep(cep_raw):
    """
    Limpa o CEP, convertendo float/int para string, removendo decimais e formatando com zeros à esquerda.
    """
    if pd.isna(cep_raw) or str(cep_raw).strip() == "":
        return ""

    cep_str = str(cep_raw).strip()

    # Se terminar em .0 (caso venha como float do pandas)
    if cep_str.endswith(".0"):
        cep_str = cep_str[:-2]

    # Remove tudo que não for dígito
    cep_str = re.sub(r'\D', '', cep_str)

    # Preenche com zero à esquerda se tiver menos de 8 dígitos
    if len(cep_str) > 0 and len(cep_str) < 8:
        cep_str = cep_str.zfill(8)

    return cep_str

def ler_planilha_excel(arquivo_excel, nome_coluna_nomes, colunas_endereco):
    """
    Lê a planilha utilizando as colunas estruturadas para o endereço.
    colunas_endereco deve ser um dicionário com as chaves:
    'logradouro', 'numero', 'bairro', 'cidade', 'cep'
    """
    try:
        if arquivo_excel.endswith('.csv'):
            df = pd.read_csv(arquivo_excel)
        else:
            df = pd.read_excel(arquivo_excel)
    except Exception as e:
        raise Exception(f"Erro ao ler o arquivo {arquivo_excel}: {e}")

    # Limpa espaços em branco ocultos nos nomes das colunas
    df.columns = df.columns.str.strip()

    # Filtra as linhas onde a coluna de nomes não é nula nem vazia
    if nome_coluna_nomes in df.columns:
        df = df[df[nome_coluna_nomes].notna() & (df[nome_coluna_nomes].astype(str).str.strip() != "")]
    else:
        # Se não tem a coluna de nome, assume índice
        df[nome_coluna_nomes] = [f"Ponto {i}" for i in range(len(df))]

    # Garante que as colunas necessárias existam no df
    for key, col in colunas_endereco.items():
        if col not in df.columns:
            df[col] = ""

    # Tratamento de nulos
    df[colunas_endereco['logradouro']] = df[colunas_endereco['logradouro']].fillna("").astype(str).str.strip()
    df[colunas_endereco['numero']] = df[colunas_endereco['numero']].fillna("").astype(str).str.strip()
    df[colunas_endereco['bairro']] = df[colunas_endereco['bairro']].fillna("").astype(str).str.strip()
    df[colunas_endereco['cidade']] = df[colunas_endereco['cidade']].fillna("").astype(str).str.strip()

    # Limpeza de CEP
    df[colunas_endereco['cep']] = df[colunas_endereco['cep']].apply(limpar_cep)

    # Formatar os números que vêm como float (ex: 196.0)
    def limpar_numero(num):
        if num.endswith('.0'): return num[:-2]
        if num.lower() == 'nan': return ""
        return num

    df[colunas_endereco['numero']] = df[colunas_endereco['numero']].apply(limpar_numero)

    # Converter para lista de dicionários para iterar
    dados_estruturados = []
    nomes = df[nome_coluna_nomes].fillna("").tolist()

    for _, row in df.iterrows():
        dados_estruturados.append({
            'logradouro': row[colunas_endereco['logradouro']],
            'numero': row[colunas_endereco['numero']],
            'bairro': row[colunas_endereco['bairro']],
            'cidade': row[colunas_endereco['cidade']],
            'cep': row[colunas_endereco['cep']]
        })

    return nomes, dados_estruturados, df

def marcar_enderecos_erro_excel(arquivo_excel, enderecos_com_erro):
    if arquivo_excel.endswith('.csv'):
        return # Não temos suporte trivial de colorir CSV

    try:
        wb = openpyxl.load_workbook(arquivo_excel)
        ws = wb.active
        fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for linha, _, _ in enderecos_com_erro:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=linha + 1, column=col).fill = fill
        wb.save(arquivo_excel)
    except Exception as e:
        pass
